"""Raw workbook -> CellRecord[].

This is the only module that imports openpyxl directly. Everything
downstream (template_analyzer, response_parser, ...) works off CellRecord,
never off openpyxl's Cell/Worksheet objects, so the rest of the codebase
stays decoupled from the Excel library's quirks.

Known limitation: openpyxl's data_only=True returns the cached result of a
formula, not a live evaluation. A workbook that was generated purely by
openpyxl (never opened in Excel) has no cached formula results, so formula
cells read back as empty. Real client workbooks -- opened, filled, and
saved in Excel -- are unaffected. Our own synthetic test files never use
formulas, so this doesn't bite Phase 13 either.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, time
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


@dataclass(frozen=True)
class CellRecord:
    sheet: str
    cell: str  # "D15"
    row: int
    col: int
    raw_value: Any  # str | int | float | datetime | date | time | bool | None
    text: str  # normalized string form ("" if empty)
    is_merged: bool
    merge_anchor: Optional[str]  # "B2" for every cell in a B2:D2 merge, including B2 itself
    number_format: str
    is_bold: bool
    has_comment: bool
    comment_text: Optional[str]
    data_validation: Optional[str]  # formula1 of any data-validation rule covering this cell


_MULTI_BLANK_LINES = re.compile(r"\n{3,}")


def _normalize_text(raw_value: Any) -> str:
    """Canonical string form of a cell's raw value.

    Only whitespace/line-ending noise is touched here -- internal commas,
    currency symbols, and slashes are preserved. Splitting on those is a
    type-aware decision that belongs to value_splitter.py (Phase 3), not here.
    """
    if raw_value is None:
        return ""
    if isinstance(raw_value, datetime):
        # datetime is a subclass of date, so this check must come first.
        if raw_value.time() == time(0, 0):
            return raw_value.date().isoformat()
        return raw_value.isoformat(sep=" ")
    if isinstance(raw_value, date):
        return raw_value.isoformat()
    if isinstance(raw_value, time):
        return raw_value.isoformat()
    if isinstance(raw_value, float) and raw_value.is_integer():
        text = str(int(raw_value))
    else:
        text = str(raw_value)

    text = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    text = _MULTI_BLANK_LINES.sub("\n\n", text)
    return text


def _build_merge_map(ws) -> tuple[dict[tuple[int, int], str], dict[str, Any]]:
    """Map every (row, col) inside any merged range -> its anchor address,
    and every anchor address -> the anchor cell's raw value. openpyxl only
    stores a value on the top-left cell of a merged range; every other cell
    in the range is a MergedCell whose .value is always None."""
    merge_map: dict[tuple[int, int], str] = {}
    anchor_raw: dict[str, Any] = {}

    for merged_range in ws.merged_cells.ranges:
        anchor_addr = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
        anchor_raw[anchor_addr] = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value

        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merge_map[(row, col)] = anchor_addr

    return merge_map, anchor_raw


def _find_data_validation(coordinate: str, dv_list: list) -> Optional[str]:
    for dv in dv_list:
        if coordinate in dv.sqref:
            return dv.formula1
    return None


def _is_nonempty(record: CellRecord) -> bool:
    return bool(record.text) or record.is_merged or record.has_comment or record.data_validation is not None


def _trim_trailing_empty(records: list[CellRecord]) -> list[CellRecord]:
    """Drop rows/columns beyond the last cell that carries any signal (text,
    merge membership, a comment, or a data-validation rule), so a sheet
    formatted out to row 1,048,576 doesn't balloon the record list."""
    last_row = 0
    last_col = 0
    for r in records:
        if _is_nonempty(r):
            last_row = max(last_row, r.row)
            last_col = max(last_col, r.col)

    if last_row == 0:
        return []

    return [r for r in records if r.row <= last_row and r.col <= last_col]


def _read_sheet(ws) -> list[CellRecord]:
    merge_map, anchor_raw = _build_merge_map(ws)
    dv_list = list(ws.data_validations.dataValidation) if ws.data_validations else []

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return []

    records: list[CellRecord] = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            key = (cell.row, cell.column)
            merge_anchor = merge_map.get(key)
            is_merged = merge_anchor is not None
            raw_value = anchor_raw[merge_anchor] if is_merged else cell.value

            comment = cell.comment
            comment_text = comment.text if comment is not None else None

            records.append(
                CellRecord(
                    sheet=ws.title,
                    cell=cell.coordinate,
                    row=cell.row,
                    col=cell.column,
                    raw_value=raw_value,
                    text=_normalize_text(raw_value),
                    is_merged=is_merged,
                    merge_anchor=merge_anchor,
                    number_format=cell.number_format or "General",
                    is_bold=bool(cell.font.bold),
                    has_comment=comment is not None,
                    comment_text=comment_text,
                    data_validation=_find_data_validation(cell.coordinate, dv_list),
                )
            )

    return _trim_trailing_empty(records)


def read_workbook(path: str) -> Workbook:
    """Thin wrapper so callers don't need to remember data_only=True."""
    return openpyxl.load_workbook(path, data_only=True)


def read_cells(path: str) -> list[CellRecord]:
    wb = read_workbook(path)
    try:
        records: list[CellRecord] = []
        for ws in wb.worksheets:
            records.extend(_read_sheet(ws))
        return records
    finally:
        wb.close()


def build_index(cells: list[CellRecord]) -> dict[tuple[str, str], CellRecord]:
    """(sheet, "D15") -> CellRecord, for O(1) lookup."""
    return {(r.sheet, r.cell): r for r in cells}
