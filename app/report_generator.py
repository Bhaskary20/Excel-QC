"""QCRun -> QC_Report.xlsx: Status, Summary, Row Analysis, Cell Analysis,
Slot Analysis, Consistency Findings. Pure presentation layer -- no QC logic
lives here.

Status is the primary, human-facing sheet -- RO | Plaza | PIU | Status |
Remarks, matching the format NHAI's own PIU-response trackers already use
(see tests/Status UPSTF_*.xlsx). The remaining sheets are the detailed
technical audit trail from BUILD_PLAN.md v2 Section 7 Phase H.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.config import Config
from app.consistency_checker import AEIE_HTMS_COLUMNS, STATUS_SEVERITY
from app.models import CellResult, QCRun, RowResult, Status

_HEADER_FILL = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_PERCENT_FORMAT = "0.0%"
_MAX_COL_WIDTH = 60

# (fill, font) per status -- light background + dark text for the "ok-ish"
# statuses (standard Excel conditional-formatting palette), a stronger dark
# red for INVALID so it reads as more severe than MISSING at a glance.
_STATUS_STYLES: dict[Status, tuple[PatternFill, Font]] = {
    Status.COMPLETE: (
        PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"),
        Font(color="FF006100"),
    ),
    Status.PARTIAL: (
        PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid"),
        Font(color="FF9C6500"),
    ),
    Status.MISSING: (
        PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),
        Font(color="FF9C0006"),
    ),
    Status.INVALID: (
        PatternFill(start_color="FFC00000", end_color="FFC00000", fill_type="solid"),
        Font(color="FFFFFFFF"),
    ),
    Status.REVIEW: (
        PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid"),
        Font(color="FF1F4E78"),
    ),
    Status.NOT_APPLICABLE: (
        PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid"),
        Font(color="FF404040"),
    ),
}

_ROW_ANALYSIS_HEADERS = [
    "S.No", "Plaza Code", "Plaza Name", "RO", "PIU", "Match Strategy",
    "Agencies (N)", "Row Status", "Completeness %", "Problem Columns", "Identity Mismatches",
]
_CELL_ANALYSIS_HEADERS = [
    "S.No", "Plaza Name", "Column", "Field Name", "Expected Count", "Detected Count",
    "Valid Count", "Invalid Count", "Missing Slots", "Completeness %", "Status", "Reason",
]
_SLOT_ANALYSIS_HEADERS = [
    "S.No", "Plaza Name", "Column", "Field Name", "Slot #", "Value", "Validation Status", "Reason",
]
_CONSISTENCY_HEADERS = ["S.No", "Plaza Name", "Column", "Kind", "Severity", "Message"]
_STATUS_SHEET_HEADERS = ["RO", "Plaza", "PIU", "Status", "Remarks"]

_NON_ISSUE_STATUSES = {Status.COMPLETE, Status.NOT_APPLICABLE}

# Plain-English field groups for the Status sheet's Remarks column, in
# template column order. K (Name) and L (Contact) are kept as separate
# entries rather than one combined "Name & Contact..." label -- they're two
# independent pieces of data, and a real client answer often gets one right
# but not the other (e.g. tests/UPSTF Case 13.08.2026.xlsx SEHATGANJ: L had
# a slot with two contact numbers that used to fail phone validation while
# K was fully correct -- the combined label would misname K as a problem
# too). Naming only the column that's actually short avoids that.
#
# N/O/P (AE/IE, HTMS/Toll Expert) are deliberately absent from this list --
# they're no longer graded by the same "any column not COMPLETE" rule as
# the others (see consistency_checker.aeie_htms_group_status), so their
# remark is built separately in _aeie_htms_remark_fragments and spliced
# into _build_remarks at the point in the loop where M's entry is
# appended, to keep template column order (M, then N/O/P, then Q).
#
# N and O (Supervision Consultant name, Team Leader name) are two people
# at the same AE/IE firm, so they're judged together as one "AE/IE" unit.
# The bar is "at least one of the two is fully covered" (every declared
# agency has a name), not just "some real data exists" -- found via real
# client data both ways:
#   - FULARA: N and P were fully covered, only O was short (2 of 7) --
#     AE/IE as a whole reads as genuinely covered, flagging it was a false
#     alarm.
#   - SEHATGANJ: N and O were each only 2 of 8 -- neither comes close to
#     full coverage, so "at least one has *some* data" was too generous a
#     bar; the remark should still name AE/IE here even though it isn't
#     literally zero.
# P (HTMS/Toll Expert) is judged on its own, still by "any real data at
# all" -- there's no second column to lean on the way N and O lean on
# each other.
_AEIE_COLUMNS = ("N", "O")
_HTMS_COLUMN = "P"
_AEIE_HTMS_REMARK_AFTER = "Address of Toll Agency"


def _aeie_htms_remark_fragments(per_column: dict[str, CellResult]) -> list[str]:
    fragments = []
    if not any(per_column[c].status == Status.COMPLETE for c in _AEIE_COLUMNS if c in per_column):
        fragments.append("AE/IE")
    if _HTMS_COLUMN in per_column and per_column[_HTMS_COLUMN].valid_count == 0:
        fragments.append("HTMS / Toll expert")
    return fragments


_FIELD_GROUPS: list[tuple[str, tuple[str, ...]]] = [
    ("Plaza Village / Location details", ("F",)),
    ("Plaza Type", ("G",)),
    ("Agency name", ("H",)),
    ("Contract details (EQ/Regular)", ("I",)),
    ("Contract Start & End dates", ("J",)),
    ("Name of Toll Manager", ("K",)),
    ("Contact details of Toll Manager", ("L",)),
    ("Address of Toll Agency", ("M",)),
    ("Traffic details", ("Q",)),
    ("Traffic details (Exemption)", ("R",)),
]

def _mask_value(value: str) -> str:
    """'9876543210' -> '98765*****' -- first 5 chars visible, rest masked."""
    if not value:
        return value
    visible = min(5, len(value))
    return value[:visible] + "*" * (len(value) - visible)


def _write_header_row(ws: Worksheet, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_style(ws: Worksheet, row_idx: int, num_cols: int, style: tuple[PatternFill, Font] | None) -> None:
    if style is None:
        return
    fill, font = style
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.font = font


def _apply_status_style(ws: Worksheet, row_idx: int, num_cols: int, status: Status) -> None:
    _apply_style(ws, row_idx, num_cols, _STATUS_STYLES.get(status))


# Reuses the same COMPLETE/PARTIAL/MISSING palette so the two sheets read
# consistently even though the Status sheet's labels are plain strings, not
# the internal Status enum.
_RESPONSE_STATUS_STYLES: dict[str, tuple[PatternFill, Font]] = {
    "Full Response": _STATUS_STYLES[Status.COMPLETE],
    "Partial Response": _STATUS_STYLES[Status.PARTIAL],
    "No Response": _STATUS_STYLES[Status.MISSING],
}


def _response_status_label(row: RowResult) -> str:
    # n_contracts == 0 means the anchor column (agency names) was never
    # filled in at all -- "No Response" regardless of whether boilerplate
    # fields like Plaza Village/Type happened to get filled in, since the
    # substantive contract data is what this QC is actually checking for.
    if row.n_contracts == 0:
        return "No Response"
    if row.status == Status.COMPLETE:
        return "Full Response"
    return "Partial Response"


def _build_remarks(row: RowResult, plaza_names_by_row: dict[int, str]) -> str:
    parts: list[str] = []
    for label, columns in _FIELD_GROUPS:
        if any(row.per_column[c].status not in _NON_ISSUE_STATUSES for c in columns if c in row.per_column):
            parts.append(label)
        if label == _AEIE_HTMS_REMARK_AFTER and all(c in row.per_column for c in AEIE_HTMS_COLUMNS):
            fragments = _aeie_htms_remark_fragments(row.per_column)
            if fragments:
                parts.append("Details of " + ", ".join(fragments))

    if row.identity_mismatches:
        parts.append("identity mismatch: " + "; ".join(row.identity_mismatches))

    if row.merged_with_rows:
        linked = [plaza_names_by_row[r] for r in row.merged_with_rows if r in plaza_names_by_row]
        if linked:
            parts.append("data shared via merged cell with " + ", ".join(linked))

    return ", ".join(parts)


def _autosize_columns(ws: Worksheet, num_cols: int) -> None:
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_length = max((len(str(c.value)) for c in ws[col_letter] if c.value is not None), default=0)
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), _MAX_COL_WIDTH)


def _finalize_sheet(ws: Worksheet, num_cols: int, num_data_rows: int, cfg: Config) -> None:
    if cfg.report.freeze_header_row:
        ws.freeze_panes = "A2"
    if cfg.report.autofilter and num_data_rows > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_data_rows + 1}"
    _autosize_columns(ws, num_cols)


def _write_status_sheet(ws: Worksheet, run: QCRun, cfg: Config) -> None:
    _write_header_row(ws, _STATUS_SHEET_HEADERS)

    plaza_names_by_row = {r.response_row: r.plaza_name for r in run.rows if r.response_row is not None}

    for row_idx, r in enumerate(run.rows, start=2):
        label = _response_status_label(r)
        ws.cell(row=row_idx, column=1, value=r.ro)
        ws.cell(row=row_idx, column=2, value=r.plaza_name)
        ws.cell(row=row_idx, column=3, value=r.piu)
        ws.cell(row=row_idx, column=4, value=label)
        ws.cell(row=row_idx, column=5, value=_build_remarks(r, plaza_names_by_row) or None)

        _apply_style(ws, row_idx, len(_STATUS_SHEET_HEADERS), _RESPONSE_STATUS_STYLES.get(label))

    _finalize_sheet(ws, len(_STATUS_SHEET_HEADERS), len(run.rows), cfg)


def _write_summary_sheet(ws: Worksheet, run: QCRun) -> None:
    s = run.workbook_summary

    ws["A1"] = "QC Report Summary"
    ws["A1"].font = Font(bold=True, size=14)

    row = 3
    for label, value in [
        ("Template", run.template_path),
        ("Response", run.response_path),
        ("Generated at", run.generated_at.isoformat(sep=" ", timespec="seconds")),
        ("Engine version", run.engine_version),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    for label, value in [
        ("Total Rows", s.total_rows),
        ("Total Cells Checked", s.total_cells_checked),
        ("Total Expected Slots", s.total_expected_slots),
        ("Total Valid Slots", s.total_valid_slots),
        ("Total Missing Slots", s.total_missing_slots),
        ("Total Invalid Slots", s.total_invalid_slots),
        ("Complete Cells", s.complete_cells),
        ("Partial Cells", s.partial_cells),
        ("Missing Cells", s.missing_cells),
        ("Invalid Cells", s.invalid_cells),
        ("Review Cells", s.review_cells),
        ("Not Applicable Cells", s.not_applicable_cells),
        ("Complete Rows", s.complete_rows),
        ("Partial Rows", s.partial_rows),
        ("Missing Rows", s.missing_rows),
        ("Invalid Rows", s.invalid_rows),
        ("Review Rows", s.review_rows),
        ("Not Applicable Rows", s.not_applicable_rows),
        ("Total Consistency Findings", s.total_consistency_findings),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    ws.cell(row=row, column=1, value="Overall Completeness %").font = Font(bold=True)
    completeness_cell = ws.cell(row=row, column=2, value=s.overall_completeness)
    if s.overall_completeness is not None:
        completeness_cell.number_format = _PERCENT_FORMAT

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40


def _write_row_analysis_sheet(ws: Worksheet, run: QCRun, cfg: Config) -> None:
    _write_header_row(ws, _ROW_ANALYSIS_HEADERS)

    for row_idx, r in enumerate(run.rows, start=2):
        problem_columns = sum(1 for cr in r.per_column.values() if cr.status not in _NON_ISSUE_STATUSES)

        ws.cell(row=row_idx, column=1, value=r.s_no)
        ws.cell(row=row_idx, column=2, value=r.plaza_code)
        ws.cell(row=row_idx, column=3, value=r.plaza_name)
        ws.cell(row=row_idx, column=4, value=r.ro)
        ws.cell(row=row_idx, column=5, value=r.piu)
        ws.cell(row=row_idx, column=6, value=r.match_strategy)
        ws.cell(row=row_idx, column=7, value=r.n_contracts)
        ws.cell(row=row_idx, column=8, value=r.status.value)
        completeness_cell = ws.cell(row=row_idx, column=9, value=r.completeness)
        if r.completeness is not None:
            completeness_cell.number_format = _PERCENT_FORMAT
        ws.cell(row=row_idx, column=10, value=problem_columns)
        ws.cell(row=row_idx, column=11, value="; ".join(r.identity_mismatches))

        _apply_status_style(ws, row_idx, len(_ROW_ANALYSIS_HEADERS), r.status)

    _finalize_sheet(ws, len(_ROW_ANALYSIS_HEADERS), len(run.rows), cfg)


def _write_cell_analysis_sheet(ws: Worksheet, run: QCRun, cfg: Config) -> None:
    _write_header_row(ws, _CELL_ANALYSIS_HEADERS)

    row_idx = 2
    for r in run.rows:
        for column_letter, cr in r.per_column.items():
            ws.cell(row=row_idx, column=1, value=r.s_no)
            ws.cell(row=row_idx, column=2, value=r.plaza_name)
            ws.cell(row=row_idx, column=3, value=column_letter)
            ws.cell(row=row_idx, column=4, value=cr.field_name)
            ws.cell(row=row_idx, column=5, value=cr.expected_count if cr.expected_count is not None else "UNKNOWN")
            ws.cell(row=row_idx, column=6, value=cr.detected_count)
            ws.cell(row=row_idx, column=7, value=cr.valid_count)
            ws.cell(row=row_idx, column=8, value=cr.invalid_count)
            ws.cell(row=row_idx, column=9, value=", ".join(str(s) for s in cr.missing_slots))
            completeness_cell = ws.cell(row=row_idx, column=10, value=cr.completeness)
            if cr.completeness is not None:
                completeness_cell.number_format = _PERCENT_FORMAT
            ws.cell(row=row_idx, column=11, value=cr.status.value)
            ws.cell(row=row_idx, column=12, value=cr.reason)

            _apply_status_style(ws, row_idx, len(_CELL_ANALYSIS_HEADERS), cr.status)
            row_idx += 1

    _finalize_sheet(ws, len(_CELL_ANALYSIS_HEADERS), row_idx - 2, cfg)


def _write_slot_analysis_sheet(ws: Worksheet, run: QCRun, cfg: Config) -> None:
    _write_header_row(ws, _SLOT_ANALYSIS_HEADERS)

    row_idx = 2
    for r in run.rows:
        for column_letter, cr in r.per_column.items():
            for sv in cr.slot_values:
                display_value = _mask_value(sv.raw) if cfg.security.redact_in_reports else sv.raw
                ws.cell(row=row_idx, column=1, value=r.s_no)
                ws.cell(row=row_idx, column=2, value=r.plaza_name)
                ws.cell(row=row_idx, column=3, value=column_letter)
                ws.cell(row=row_idx, column=4, value=cr.field_name)
                ws.cell(row=row_idx, column=5, value=sv.slot)
                ws.cell(row=row_idx, column=6, value=display_value)
                ws.cell(row=row_idx, column=7, value="VALID" if sv.verdict.is_valid else "INVALID")
                ws.cell(row=row_idx, column=8, value=sv.verdict.reason)
                row_idx += 1

    _finalize_sheet(ws, len(_SLOT_ANALYSIS_HEADERS), row_idx - 2, cfg)


def _write_consistency_findings_sheet(ws: Worksheet, run: QCRun, cfg: Config) -> None:
    _write_header_row(ws, _CONSISTENCY_HEADERS)

    entries: list[tuple[RowResult, object]] = [
        (r, finding) for r in run.rows for finding in r.consistency_findings
    ]
    entries.sort(key=lambda pair: STATUS_SEVERITY[pair[1].severity], reverse=True)

    for row_idx, (r, finding) in enumerate(entries, start=2):
        ws.cell(row=row_idx, column=1, value=r.s_no)
        ws.cell(row=row_idx, column=2, value=r.plaza_name)
        ws.cell(row=row_idx, column=3, value=finding.column or "")
        ws.cell(row=row_idx, column=4, value=finding.kind)
        ws.cell(row=row_idx, column=5, value=finding.severity.value)
        ws.cell(row=row_idx, column=6, value=finding.message)

        _apply_status_style(ws, row_idx, len(_CONSISTENCY_HEADERS), finding.severity)

    _finalize_sheet(ws, len(_CONSISTENCY_HEADERS), len(entries), cfg)


def generate_report(run: QCRun, output_path: str, cfg: Config) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_status_sheet(wb.create_sheet("Status"), run, cfg)
    _write_summary_sheet(wb.create_sheet("Summary"), run)
    _write_row_analysis_sheet(wb.create_sheet("Row Analysis"), run, cfg)
    _write_cell_analysis_sheet(wb.create_sheet("Cell Analysis"), run, cfg)
    _write_slot_analysis_sheet(wb.create_sheet("Slot Analysis"), run, cfg)
    _write_consistency_findings_sheet(wb.create_sheet("Consistency Findings"), run, cfg)

    output_dir = Path(output_path).parent
    if output_dir and not output_dir.exists():
        output_dir.mkdir(parents=True, exist_ok=True)

    wb.save(output_path)
