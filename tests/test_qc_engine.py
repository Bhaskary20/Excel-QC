"""Phase G gate: an untouched-scaffold workbook scores 0% (not 100%), a
fully-correct workbook scores 100%, and workbook-level aggregation
(Section 5.3) is arithmetically correct -- including NOT_APPLICABLE cells
never dragging completeness down despite carrying an expected_count.
"""

from pathlib import Path

import openpyxl
import pytest

import app.validators  # noqa: F401 -- real validators, not the placeholder
from app.config import load_config
from app.models import Status
from app.qc_engine import run_qc, run_qc_from_cells
from app.template_spec import COLUMNS, FIRST_DATA_ROW, HEADER_ROW, SHEET_NAME

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "template" / "Format.xlsx"


@pytest.fixture(scope="module")
def cfg():
    return load_config()


_FULL_ROW = {
    "A": "1", "B": "111111", "C": "ALPHA", "D": "Delhi", "E": "PIU-A",
    "F": "Alpha Village, Chainage 12+300, Alpha City, 462001",
    "G": "BOT",
    "H": "1. Agency One\n2. Agency Two",
    "I": "1. EQ (3 months)\n2. Regular (1 year)",
    "J": "1. 10/08/2021 - 10/11/2021\n2. 10/11/2021 - 14/01/2026",
    "K": "1. Rahul Sharma\n2. Amit Kumar",
    "L": "1. 9876543210\n2. 9876543211",
    "M": "1. Toll Plaza Road, Bhopal\n2. Toll Plaza Road, Bhopal",
    "N": "1. Consultant A\n2. Consultant B",
    "O": "1. Team Lead A\n2. Team Lead B",
    "P": "1. HTMS A\n2. HTMS B",
    "Q": "1. 1500\n2. 1600",
    "R": "1. 100\n2. 110",
    "S": "all good",
}
_BLANK_INPUT_COLUMNS = {k: "" for k in _FULL_ROW if k not in ("A", "B", "C", "D", "E")}


def _write_workbook(path: Path, rows: list[dict[str, str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    for letter, spec in COLUMNS.items():
        ws.cell(row=HEADER_ROW, column=spec.index, value=spec.header)
    for i, row_values in enumerate(rows):
        row_num = FIRST_DATA_ROW + i
        for letter, text in row_values.items():
            ws.cell(row=row_num, column=COLUMNS[letter].index, value=text)
    wb.save(path)


def _full_row(**overrides) -> dict[str, str]:
    values = dict(_FULL_ROW)
    values.update(overrides)
    return values


# ============================================================================
# The 2 Done-when examples
# ============================================================================


def test_untouched_template_as_response_scores_zero_percent(cfg):
    run = run_qc(str(TEMPLATE_PATH), str(TEMPLATE_PATH), cfg)

    assert run.workbook_summary.total_rows == 115
    assert run.workbook_summary.overall_completeness == 0.0
    assert run.workbook_summary.total_valid_slots == 0
    assert run.workbook_summary.missing_rows == 115
    assert run.workbook_summary.complete_rows == 0


def test_fully_correct_small_workbook_scores_100_percent(cfg, tmp_path):
    template_path = tmp_path / "template.xlsx"
    response_path = tmp_path / "response.xlsx"
    # Template: identity columns filled, everything else blank (a minimal
    # synthetic template, not the real 115-row one).
    _write_workbook(template_path, [_full_row(**_BLANK_INPUT_COLUMNS)])
    _write_workbook(response_path, [_full_row()])

    run = run_qc(str(template_path), str(response_path), cfg)

    assert run.workbook_summary.overall_completeness == 1.0
    assert run.workbook_summary.complete_rows == 1
    assert run.workbook_summary.missing_cells == 0
    assert run.workbook_summary.invalid_cells == 0


# ============================================================================
# Aggregation arithmetic across a mixed multi-row workbook
# ============================================================================


def test_aggregation_across_mixed_rows(cfg, tmp_path):
    template_path = tmp_path / "template.xlsx"
    response_path = tmp_path / "response.xlsx"

    row1 = _full_row(A="1", B="111111", C="ALPHA")  # fully correct
    row2 = _full_row(A="2", B="222222", C="BETA", L="9876543210")  # only 1 of 2 phones -> PARTIAL
    row3 = _full_row(A="3", B="333333", C="GAMMA", **_BLANK_INPUT_COLUMNS)  # untouched -> MISSING

    template_rows = [
        _full_row(A="1", B="111111", C="ALPHA", **_BLANK_INPUT_COLUMNS),
        _full_row(A="2", B="222222", C="BETA", **_BLANK_INPUT_COLUMNS),
        _full_row(A="3", B="333333", C="GAMMA", **_BLANK_INPUT_COLUMNS),
    ]
    _write_workbook(template_path, template_rows)
    _write_workbook(response_path, [row1, row2, row3])

    run = run_qc(str(template_path), str(response_path), cfg)

    assert run.workbook_summary.total_rows == 3
    assert run.workbook_summary.complete_rows == 1
    assert run.workbook_summary.partial_rows == 1
    assert run.workbook_summary.missing_rows == 1

    by_name = {r.plaza_name: r for r in run.rows}
    assert by_name["ALPHA"].status == Status.COMPLETE
    assert by_name["BETA"].status == Status.PARTIAL
    assert by_name["GAMMA"].status == Status.MISSING


# ============================================================================
# NOT_APPLICABLE must never drag completeness down (regression: caught by
# hand-tracing the untouched-scaffold smoke test before this suite existed)
# ============================================================================


def test_not_applicable_excluded_from_completeness_denominator(cfg, tmp_path):
    template_path = tmp_path / "template.xlsx"
    response_path = tmp_path / "response.xlsx"
    _write_workbook(template_path, [_full_row(**_BLANK_INPUT_COLUMNS)])
    # Everything filled correctly EXCEPT Remarks (S), left blank -- optional,
    # so it must read NOT_APPLICABLE and not affect the 100% outcome.
    _write_workbook(response_path, [_full_row(S="")])

    run = run_qc(str(template_path), str(response_path), cfg)

    row = run.rows[0]
    assert row.per_column["S"].status == Status.NOT_APPLICABLE
    assert run.workbook_summary.overall_completeness == 1.0
    assert run.workbook_summary.not_applicable_cells == 1


# ============================================================================
# extra_response_rows: bounded to the data range, no false positives from
# the title/header rows (regression: caught in manual smoke testing)
# ============================================================================


def test_extra_response_rows_excludes_title_and_header(cfg):
    run = run_qc(str(TEMPLATE_PATH), str(TEMPLATE_PATH), cfg)
    assert run.extra_response_rows == []


def test_extra_response_rows_detects_a_genuinely_unmatched_plaza(cfg, tmp_path):
    template_path = tmp_path / "template.xlsx"
    response_path = tmp_path / "response.xlsx"
    _write_workbook(template_path, [_full_row(**_BLANK_INPUT_COLUMNS)])
    # Response has the matched row PLUS an extra plaza the template doesn't know about.
    extra_row = _full_row(A="99", B="999999", C="UNKNOWN PLAZA")
    _write_workbook(response_path, [_full_row(), extra_row])

    run = run_qc(str(template_path), str(response_path), cfg)

    assert len(run.extra_response_rows) == 1
    assert "row 4" in run.extra_response_rows[0]  # FIRST_DATA_ROW(3) + 1


# ============================================================================
# run_qc_from_cells works without real files (used throughout Phases E/F)
# ============================================================================


def test_run_qc_from_cells_usable_without_files(cfg):
    from app.excel_reader import read_cells

    cells = read_cells(str(TEMPLATE_PATH))
    run = run_qc_from_cells(cells, cells, cfg)
    assert run.workbook_summary.total_rows == 115
    assert run.template_path == ""
    assert run.response_path == ""
