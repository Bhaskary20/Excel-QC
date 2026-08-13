"""Phase H gate: all six sheets present with exact headers, the file
round-trips through openpyxl cleanly, completeness is a real number (not
a string), and the report generates correctly against the real template
end to end."""

import dataclasses
from pathlib import Path

import openpyxl
import pytest

import app.validators  # noqa: F401
from app.config import load_config
from app.models import (
    CellResult,
    ConsistencyFinding,
    QCRun,
    RowResult,
    SheetSummary,
    SlotValue,
    Status,
    ValueVerdict,
    WorkbookSummary,
)
from app.qc_engine import run_qc
from app.report_generator import generate_report
from app.template_spec import ValueType

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "template" / "Format.xlsx"


@pytest.fixture(scope="module")
def cfg():
    return load_config()


def _cell_result(
    column, field_name, value_type, expected, detected, valid, invalid, missing_slots, status,
    completeness=None, slot_values=None,
) -> CellResult:
    return CellResult(
        sheet="Data Sheet for UP STF", cell=f"{column}15", column=column, field_name=field_name,
        value_type=value_type, expected_count=expected, detected_count=detected, valid_count=valid,
        invalid_count=invalid, missing_count=(max(expected - valid, 0) if expected is not None else None),
        missing_slots=missing_slots, completeness=completeness, status=status, confidence=1.0,
        reason="test reason", slot_values=slot_values or [],
    )


def _sample_run() -> QCRun:
    l_values = [
        SlotValue(slot=1, raw="9876543210", verdict=ValueVerdict(is_valid=True, normalized="9876543210")),
        SlotValue(slot=2, raw="9876", verdict=ValueVerdict(is_valid=False, normalized=None, reason="expected 10 digits, got 4")),
    ]
    per_column_partial = {
        "H": _cell_result("H", "Agency name", ValueType.TEXT, None, 4, 4, 0, [], Status.COMPLETE),
        "K": _cell_result("K", "Name of Toll Plaza Manager", ValueType.NAME, 4, 4, 4, 0, [], Status.COMPLETE),
        "L": _cell_result(
            "L", "Contact of Toll Manager", ValueType.PHONE, 4, 2, 1, 1, [3, 4], Status.PARTIAL,
            completeness=0.25, slot_values=l_values,
        ),
    }
    per_column_complete = {
        "H": _cell_result("H", "Agency name", ValueType.TEXT, None, 2, 2, 0, [], Status.COMPLETE),
    }

    rows = [
        RowResult(
            sheet="Data Sheet for UP STF", response_row=15, s_no=13, plaza_code="345061",
            plaza_name="SEHATGANJ", ro="Bhopal", piu="Bhopal", match_strategy="s_no", n_contracts=4,
            per_column=per_column_partial,
            consistency_findings=[
                ConsistencyFinding(kind="slot_count_mismatch", column="L", severity=Status.REVIEW, message="4 agencies declared; missing at slots 3, 4"),
            ],
            status=Status.PARTIAL, completeness=0.25,
        ),
        RowResult(
            sheet="Data Sheet for UP STF", response_row=16, s_no=14, plaza_code="222222",
            plaza_name="MAUHARI", ro="Bhopal", piu="Chhatarpur", match_strategy="s_no", n_contracts=2,
            per_column=per_column_complete, consistency_findings=[],
            status=Status.COMPLETE, completeness=1.0,
        ),
    ]
    return QCRun(
        template_path="template.xlsx", response_path="response.xlsx",
        rows=rows,
        sheet_summary=SheetSummary(sheet="Data Sheet for UP STF", total_rows=2, expected_slots=8, valid_slots=6, completeness=0.75),
        workbook_summary=WorkbookSummary(
            total_rows=2, total_cells_checked=3, total_expected_slots=8, total_valid_slots=6,
            complete_cells=2, partial_cells=1, overall_completeness=0.75, total_consistency_findings=1,
        ),
        extra_response_rows=[],
    )


def _generate(run, cfg, tmp_path, name="QC_Report.xlsx"):
    output = tmp_path / name
    generate_report(run, str(output), cfg)
    return output


# ============================================================================
# Structure: all five sheets, exact headers
# ============================================================================


def test_report_generates_all_six_sheets_in_order(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    wb = openpyxl.load_workbook(output)
    assert wb.sheetnames == [
        "Status", "Summary", "Row Analysis", "Cell Analysis", "Slot Analysis", "Consistency Findings",
    ]


def test_status_sheet_headers_exact(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Status"]
    headers = [c.value for c in ws[1]]
    assert headers == ["RO", "Plaza", "PIU", "Status", "Remarks"]


def test_row_analysis_headers_exact(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Row Analysis"]
    headers = [c.value for c in ws[1]]
    assert headers == [
        "S.No", "Plaza Code", "Plaza Name", "RO", "PIU", "Match Strategy",
        "Agencies (N)", "Row Status", "Completeness %", "Problem Columns", "Identity Mismatches",
    ]


def test_cell_analysis_headers_exact(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Cell Analysis"]
    headers = [c.value for c in ws[1]]
    assert headers == [
        "S.No", "Plaza Name", "Column", "Field Name", "Expected Count", "Detected Count",
        "Valid Count", "Invalid Count", "Missing Slots", "Completeness %", "Status", "Reason",
    ]


def test_slot_analysis_headers_exact(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Slot Analysis"]
    headers = [c.value for c in ws[1]]
    assert headers == ["S.No", "Plaza Name", "Column", "Field Name", "Slot #", "Value", "Validation Status", "Reason"]


def test_consistency_findings_headers_exact(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Consistency Findings"]
    headers = [c.value for c in ws[1]]
    assert headers == ["S.No", "Plaza Name", "Column", "Kind", "Severity", "Message"]


# ============================================================================
# Status sheet: RO | Plaza | PIU | Status | Remarks, matching the format of
# tests/Status UPSTF_*.xlsx (see report_generator.py's module docstring).
# ============================================================================


def test_status_sheet_full_response_for_a_fully_complete_row(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Status"]
    # MAUHARI (row 3): status=COMPLETE, n_contracts=2 -> Full Response, no remarks.
    assert ws.cell(row=3, column=1).value == "Bhopal"
    assert ws.cell(row=3, column=2).value == "MAUHARI"
    assert ws.cell(row=3, column=3).value == "Chhatarpur"
    assert ws.cell(row=3, column=4).value == "Full Response"
    assert ws.cell(row=3, column=5).value is None


def test_status_sheet_partial_response_and_field_group_remark(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Status"]
    # SEHATGANJ (row 2): H and K complete, L partial -- Partial Response, and
    # the remark names only "Contact details of Toll Manager" (L), not K --
    # K and L are graded and reported independently, not as one combined
    # "Name & Contact..." label that would misname K as a problem too.
    assert ws.cell(row=2, column=4).value == "Partial Response"
    assert ws.cell(row=2, column=5).value == "Contact details of Toll Manager"


def test_status_sheet_no_response_when_n_contracts_zero(cfg, tmp_path):
    run = _sample_run()
    run.rows[0] = dataclasses.replace(run.rows[0], n_contracts=0)
    output = _generate(run, cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Status"]
    assert ws.cell(row=2, column=4).value == "No Response"


def test_status_sheet_slot_count_mismatch_not_double_reported(cfg, tmp_path):
    # SEHATGANJ's one consistency finding is a slot_count_mismatch on L --
    # that's the same gap the "Contact details of Toll Manager" field-group
    # remark already names, so it must not also appear as its own remarks entry.
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Status"]
    remarks = ws.cell(row=2, column=5).value
    assert remarks.count(",") == 0  # exactly one item, no second clause appended


def test_status_sheet_identity_mismatch_appended_to_remarks(cfg, tmp_path):
    run = _sample_run()
    run.rows[1] = dataclasses.replace(
        run.rows[1], identity_mismatches=["Plaza Code: template=345061, response=999999"]
    )
    output = _generate(run, cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Status"]
    assert "identity mismatch" in ws.cell(row=3, column=5).value


def test_status_sheet_aeie_htms_remark_fires_when_group_has_no_data(cfg, tmp_path):
    # N/O/P are graded as a group (see consistency_checker.aeie_htms_group_status),
    # not per-column -- when none of the three has any real data, the
    # "Details of AE/IE, HTMS / Toll expert" remark must still appear, same
    # as it always has.
    run = _sample_run()
    bad_n_o_p = {
        "N": _cell_result("N", "Supervision Consultant (AE / IE) name and its address", ValueType.NAME, 4, 0, 0, 0, [1, 2, 3, 4], Status.MISSING),
        "O": _cell_result("O", "Team Leader name (AE/ IE ) during the said Contract Period", ValueType.NAME, 4, 0, 0, 0, [1, 2, 3, 4], Status.MISSING),
        "P": _cell_result("P", "HTMS / Toll Expert during said contract period", ValueType.TEXT, 4, 0, 0, 0, [1, 2, 3, 4], Status.MISSING),
    }
    run.rows[1] = dataclasses.replace(
        run.rows[1], per_column={**run.rows[1].per_column, **bad_n_o_p}, status=Status.MISSING,
    )
    output = _generate(run, cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Status"]
    assert ws.cell(row=3, column=4).value == "Partial Response"
    remarks = ws.cell(row=3, column=5).value
    assert "Details of AE/IE, HTMS / Toll expert" in remarks


def test_status_sheet_aeie_htms_remark_absent_when_group_status_is_complete(cfg, tmp_path):
    # O is only PARTIAL on its own (2 of 4), but N and P each have real data
    # too, so the AE/IE/HTMS *group* is COMPLETE. The remark must go by the
    # group verdict, not fire just because one of the three individually
    # falls short of the agency count.
    run = _sample_run()
    n_o_p = {
        "N": _cell_result("N", "Supervision Consultant (AE / IE) name and its address", ValueType.NAME, 4, 4, 4, 0, [], Status.COMPLETE),
        "O": _cell_result("O", "Team Leader name (AE/ IE ) during the said Contract Period", ValueType.NAME, 4, 2, 2, 0, [3, 4], Status.PARTIAL),
        "P": _cell_result("P", "HTMS / Toll Expert during said contract period", ValueType.TEXT, 4, 4, 4, 0, [], Status.COMPLETE),
    }
    run.rows[1] = dataclasses.replace(run.rows[1], per_column={**run.rows[1].per_column, **n_o_p})
    output = _generate(run, cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Status"]
    remarks = ws.cell(row=3, column=5).value
    assert remarks is None or "AE/IE" not in remarks


def test_status_sheet_aeie_htms_remark_fires_when_neither_is_fully_covered(cfg, tmp_path):
    # Real case (SEHATGANJ, tests/UPSTF Case 13.08.2026.xlsx): N and O each
    # have only 2 of 8 agencies covered -- neither is close to complete,
    # even though it isn't literally zero. Unlike the FULARA case (one of
    # the two fully covers every agency), "has any data" is too generous a
    # bar here -- the remark must still name AE/IE, on top of HTMS (P,
    # whole-cell "NOT ASSIGNED", zero).
    run = _sample_run()
    n_o_p = {
        "N": _cell_result("N", "Supervision Consultant (AE / IE) name and its address", ValueType.NAME, 8, 2, 2, 0, [3, 4, 5, 6, 7, 8], Status.PARTIAL),
        "O": _cell_result("O", "Team Leader name (AE/ IE ) during the said Contract Period", ValueType.NAME, 8, 2, 2, 0, [3, 4, 5, 6, 7, 8], Status.PARTIAL),
        "P": _cell_result("P", "HTMS / Toll Expert during said contract period", ValueType.TEXT, 8, 0, 0, 0, list(range(1, 9)), Status.MISSING),
    }
    run.rows[1] = dataclasses.replace(run.rows[1], per_column={**run.rows[1].per_column, **n_o_p})
    output = _generate(run, cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Status"]
    remarks = ws.cell(row=3, column=5).value
    assert remarks == "Details of AE/IE, HTMS / Toll expert"


def test_status_sheet_row_style_reflects_response_label(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Status"]
    complete_fill = ws.cell(row=3, column=1).fill  # MAUHARI, Full Response
    assert complete_fill.start_color.rgb == "FFC6EFCE"


# ============================================================================
# Content correctness
# ============================================================================


def test_row_analysis_problem_columns_excludes_complete_and_na(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Row Analysis"]
    # SEHATGANJ row: H is COMPLETE, L is PARTIAL -- 1 problem column.
    assert ws.cell(row=2, column=10).value == 1
    # MAUHARI row: only H, which is COMPLETE -- 0 problem columns.
    assert ws.cell(row=3, column=10).value == 0


def test_cell_analysis_missing_slots_joined_as_string(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Cell Analysis"]
    rows_by_column = {ws.cell(row=r, column=3).value: r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value == 13}
    l_row = rows_by_column["L"]
    assert ws.cell(row=l_row, column=9).value == "3, 4"


def test_cell_analysis_unknown_expected_count_shown_not_blank(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Cell Analysis"]
    # H's expected_count is None (anchor column) in this fixture.
    h_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=3).value == "H"]
    assert all(ws.cell(row=r, column=5).value == "UNKNOWN" for r in h_rows)


def test_slot_analysis_one_row_per_slot_value(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Slot Analysis"]
    # Only L has slot_values in the fixture: 2 entries.
    assert ws.max_row == 3  # header + 2


def test_slot_analysis_validation_status_strings(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Slot Analysis"]
    statuses = [ws.cell(row=r, column=7).value for r in range(2, ws.max_row + 1)]
    assert statuses == ["VALID", "INVALID"]


def test_consistency_findings_content(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Consistency Findings"]
    assert ws.max_row == 2  # header + 1 finding
    assert ws.cell(row=2, column=4).value == "slot_count_mismatch"
    assert ws.cell(row=2, column=5).value == "REVIEW"


def test_summary_sheet_has_key_metrics(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Summary"]
    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "Total Rows" in labels
    assert "Overall Completeness %" in labels
    assert "Template" in labels


# ============================================================================
# Completeness is numeric (sorts correctly), not a string
# ============================================================================


def test_completeness_is_numeric_with_percent_format(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Row Analysis"]
    cell = ws.cell(row=2, column=9)  # SEHATGANJ, completeness=0.25
    assert isinstance(cell.value, float)
    assert cell.value == pytest.approx(0.25)
    assert cell.number_format == "0.0%"


# ============================================================================
# Redaction
# ============================================================================


def test_no_redaction_by_default(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Slot Analysis"]
    values = [ws.cell(row=r, column=6).value for r in range(2, ws.max_row + 1)]
    assert "9876543210" in values


def test_redaction_masks_values_when_enabled(tmp_path):
    cfg = load_config(overrides={"security": {"redact_in_reports": True}})
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Slot Analysis"]
    values = [ws.cell(row=r, column=6).value for r in range(2, ws.max_row + 1)]
    assert "9876543210" not in values
    assert "98765*****" in values


# ============================================================================
# Formatting
# ============================================================================


def test_status_fill_applied_to_complete_row(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Row Analysis"]
    fill = ws.cell(row=3, column=1).fill  # MAUHARI row, status=COMPLETE
    assert fill.start_color.rgb == "FFC6EFCE"


def test_freeze_panes_and_autofilter_set(cfg, tmp_path):
    output = _generate(_sample_run(), cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Cell Analysis"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


def test_output_directory_created_if_missing(cfg, tmp_path):
    output = tmp_path / "nested" / "dir" / "QC_Report.xlsx"
    generate_report(_sample_run(), str(output), cfg)
    assert output.exists()


def test_empty_run_does_not_crash(cfg, tmp_path):
    empty_run = QCRun(
        template_path="t.xlsx", response_path="r.xlsx",
        sheet_summary=SheetSummary(sheet="Data Sheet for UP STF"),
        workbook_summary=WorkbookSummary(),
    )
    output = _generate(empty_run, cfg, tmp_path, name="empty.xlsx")
    wb = openpyxl.load_workbook(output)
    assert wb.sheetnames == [
        "Status", "Summary", "Row Analysis", "Cell Analysis", "Slot Analysis", "Consistency Findings",
    ]


# ============================================================================
# Consistency Findings sorted most-severe-first
# ============================================================================


def test_findings_sorted_most_severe_first(cfg, tmp_path):
    rows = [
        RowResult(
            sheet="s", response_row=1, s_no=1, plaza_code="1", plaza_name="A", ro="R", piu="P",
            match_strategy="s_no", n_contracts=1, per_column={},
            consistency_findings=[
                ConsistencyFinding(kind="date_window_coverage", column="J", severity=Status.REVIEW, message="review-level"),
            ],
            status=Status.REVIEW, completeness=None,
        ),
        RowResult(
            sheet="s", response_row=2, s_no=2, plaza_code="2", plaza_name="B", ro="R", piu="P",
            match_strategy="s_no", n_contracts=1, per_column={},
            consistency_findings=[
                ConsistencyFinding(kind="hypothetical_invalid_finding", column="L", severity=Status.INVALID, message="invalid-level"),
            ],
            status=Status.INVALID, completeness=None,
        ),
    ]
    run = QCRun(
        template_path="t.xlsx", response_path="r.xlsx", rows=rows,
        sheet_summary=SheetSummary(sheet="s"), workbook_summary=WorkbookSummary(),
    )
    output = _generate(run, cfg, tmp_path)
    ws = openpyxl.load_workbook(output)["Consistency Findings"]
    severities = [ws.cell(row=r, column=5).value for r in range(2, ws.max_row + 1)]
    assert severities == ["INVALID", "REVIEW"]  # most severe first


# ============================================================================
# End-to-end against the real template (Done-when: opens without repair)
# ============================================================================


def test_end_to_end_against_real_template_untouched(cfg, tmp_path):
    run = run_qc(str(TEMPLATE_PATH), str(TEMPLATE_PATH), cfg)
    output = _generate(run, cfg, tmp_path, name="real_report.xlsx")

    wb = openpyxl.load_workbook(output)
    assert wb.sheetnames == [
        "Status", "Summary", "Row Analysis", "Cell Analysis", "Slot Analysis", "Consistency Findings",
    ]
    assert wb["Row Analysis"].max_row == 116  # header + 115 plazas
    assert wb["Cell Analysis"].max_row == 1 + 115 * 14

    # Untouched template as its own response -> every row has n_contracts=0,
    # so the Status sheet's central "did they even respond" signal must
    # read "No Response" across the board, with no exceptions.
    status_ws = wb["Status"]
    assert status_ws.max_row == 116
    labels = {status_ws.cell(row=r, column=4).value for r in range(2, status_ws.max_row + 1)}
    assert labels == {"No Response"}
