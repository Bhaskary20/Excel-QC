"""Phase I gate: end-to-end run on synthetic responses, and every bad-input
case (missing file, .xls, wrong sheet, bad config path) produces a clear
message and exit code 2 -- never a raw traceback.
"""

import json
from pathlib import Path

import openpyxl
import pytest

from main import DEFAULT_TEMPLATE, main
from app.template_spec import COLUMNS, FIRST_DATA_ROW, HEADER_ROW, SHEET_NAME

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "template" / "Format.xlsx"

_FULL_ROW = {
    "A": "1", "B": "111111", "C": "ALPHA", "D": "Delhi", "E": "PIU-A",
    "F": "Alpha Village, Chainage 12+300, Alpha City, 462001",
    "G": "BOT",
    "H": "1. Agency One\n2. Agency Two",
    "I": "1. EQ (3 months)\n2. Regular (1 year)",
    "J": "1. 10/08/2021 - 10/11/2021\n2. 10/11/2021 - 14/01/2026",
    "K": "1. Rahul Sharma\n2. Amit Kumar",
    "L": "1. 9876543210\n2. 9876543211",
    "M": "1. Toll Plaza Road, Bhopal\n2. Toll Plaza Road, Bhopal",
    "N": "1. Consultant A\n2. Consultant B",
    "O": "1. Team Lead A\n2. Team Lead B",
    "P": "1. HTMS A\n2. HTMS B",
    "Q": "1. 1500\n2. 1600",
    "R": "1. 100\n2. 110",
    "S": "all good",
}
_BLANK_INPUT_COLUMNS = {k: "" for k in _FULL_ROW if k not in ("A", "B", "C", "D", "E")}


def _write_workbook(path: Path, rows: list[dict[str, str]], sheet_name: str = SHEET_NAME) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for letter, spec in COLUMNS.items():
        ws.cell(row=HEADER_ROW, column=spec.index, value=spec.header)
    for i, row_values in enumerate(rows):
        row_num = FIRST_DATA_ROW + i
        for letter, text in row_values.items():
            ws.cell(row=row_num, column=COLUMNS[letter].index, value=text)
    wb.save(path)


def _full_row(**overrides) -> dict[str, str]:
    values = dict(_FULL_ROW)
    values.update(overrides)
    return values


@pytest.fixture()
def mini_template(tmp_path) -> Path:
    path = tmp_path / "template.xlsx"
    _write_workbook(path, [_full_row(**_BLANK_INPUT_COLUMNS)])
    return path


# ============================================================================
# Successful end-to-end runs
# ============================================================================


def test_successful_run_exits_zero_and_writes_report(mini_template, tmp_path, capsys):
    response_path = tmp_path / "response.xlsx"
    _write_workbook(response_path, [_full_row()])
    output_path = tmp_path / "QC_Report.xlsx"

    exit_code = main(["--template", str(mini_template), "--response", str(response_path), "--output", str(output_path)])

    assert exit_code == 0
    assert output_path.exists()
    out = capsys.readouterr().out
    assert "Overall completeness: 100.00%" in out


def test_quiet_mode_prints_one_line(mini_template, tmp_path, capsys):
    response_path = tmp_path / "response.xlsx"
    _write_workbook(response_path, [_full_row()])
    output_path = tmp_path / "QC_Report.xlsx"

    main(["--template", str(mini_template), "--response", str(response_path), "--output", str(output_path), "--quiet"])

    out = capsys.readouterr().out.strip()
    assert len(out.splitlines()) == 1
    assert "100.00%" in out


def test_default_template_points_to_the_real_committed_file():
    assert Path(DEFAULT_TEMPLATE).exists()
    assert Path(DEFAULT_TEMPLATE).resolve() == TEMPLATE_PATH.resolve()


# ============================================================================
# --strict
# ============================================================================


def test_strict_fails_on_incomplete_response(mini_template, tmp_path):
    response_path = tmp_path / "response.xlsx"
    _write_workbook(response_path, [_full_row(**_BLANK_INPUT_COLUMNS)])  # untouched -> MISSING
    output_path = tmp_path / "QC_Report.xlsx"

    exit_code = main(["--template", str(mini_template), "--response", str(response_path), "--output", str(output_path), "--strict"])

    assert exit_code == 1


def test_strict_succeeds_on_fully_complete_response(mini_template, tmp_path):
    response_path = tmp_path / "response.xlsx"
    _write_workbook(response_path, [_full_row()])
    output_path = tmp_path / "QC_Report.xlsx"

    exit_code = main(["--template", str(mini_template), "--response", str(response_path), "--output", str(output_path), "--strict"])

    assert exit_code == 0


# ============================================================================
# Bad input: clear errors, exit code 2, never a traceback
# ============================================================================


def test_missing_response_file_exits_2(mini_template, tmp_path, capsys):
    exit_code = main(["--template", str(mini_template), "--response", str(tmp_path / "does_not_exist.xlsx")])
    assert exit_code == 2
    err = capsys.readouterr().err
    assert "not found" in err
    assert "Traceback" not in err


def test_xls_extension_exits_2(mini_template, tmp_path, capsys):
    fake_xls = tmp_path / "response.xls"
    fake_xls.write_text("not a real xls file")

    exit_code = main(["--template", str(mini_template), "--response", str(fake_xls)])

    assert exit_code == 2
    err = capsys.readouterr().err
    assert ".xls" in err
    assert "Traceback" not in err


def test_corrupted_xlsx_exits_2_not_a_traceback(mini_template, tmp_path, capsys):
    corrupted = tmp_path / "corrupted.xlsx"
    corrupted.write_bytes(b"this is not a real zip/xlsx file at all")

    exit_code = main(["--template", str(mini_template), "--response", str(corrupted)])

    assert exit_code == 2
    err = capsys.readouterr().err
    assert "Traceback" not in err


def test_wrong_sheet_name_exits_2_with_sheet_list(mini_template, tmp_path, capsys):
    response_path = tmp_path / "response.xlsx"
    _write_workbook(response_path, [_full_row()], sheet_name="Some Other Sheet")
    # Add a second sheet WITH content -- an empty sheet gets trimmed away
    # entirely by excel_reader, which would leave only one real sheet name
    # and let resolve_sheet's "only one sheet" fallback mask the ambiguity
    # this test is meant to exercise.
    wb = openpyxl.load_workbook(response_path)
    extra = wb.create_sheet("Yet Another Sheet")
    extra["A1"] = "unrelated content"
    wb.save(response_path)

    exit_code = main(["--template", str(mini_template), "--response", str(response_path)])

    assert exit_code == 2
    err = capsys.readouterr().err
    assert "Some Other Sheet" in err


def test_nonexistent_config_path_exits_2(mini_template, tmp_path, capsys):
    response_path = tmp_path / "response.xlsx"
    _write_workbook(response_path, [_full_row()])

    exit_code = main([
        "--template", str(mini_template), "--response", str(response_path),
        "--config", str(tmp_path / "no_such_config.yaml"),
    ])

    assert exit_code == 2
    err = capsys.readouterr().err
    assert "Traceback" not in err


# ============================================================================
# Identical template/response: a warning, not an error
# ============================================================================


def test_identical_template_and_response_warns_but_succeeds(capsys, tmp_path):
    output_path = tmp_path / "QC_Report.xlsx"
    exit_code = main(["--response", str(TEMPLATE_PATH), "--template", str(TEMPLATE_PATH), "--output", str(output_path)])

    assert exit_code == 0
    err = capsys.readouterr().err
    assert "identical to the template" in err


# ============================================================================
# --dump-json
# ============================================================================


def test_dump_json_writes_valid_json(mini_template, tmp_path):
    response_path = tmp_path / "response.xlsx"
    _write_workbook(response_path, [_full_row()])
    output_path = tmp_path / "QC_Report.xlsx"
    json_path = tmp_path / "dump.json"

    main([
        "--template", str(mini_template), "--response", str(response_path),
        "--output", str(output_path), "--dump-json", str(json_path),
    ])

    assert json_path.exists()
    data = json.loads(json_path.read_text(encoding="utf-8"))
    assert data["workbook_summary"]["overall_completeness"] == 1.0
    assert len(data["rows"]) == 1
    assert data["rows"][0]["status"] == "COMPLETE"
