"""Phase A gate: the tripwire. If template/Format.xlsx is ever silently
edited -- a header reworded, a scaffold string changed, rows added or
removed -- these tests must fail loudly, because every later phase trusts
template_spec.py without re-reading the file itself.
"""

from pathlib import Path

import openpyxl
import pytest

from app.template_spec import (
    ANCHOR_COLUMN,
    COLUMN_LETTERS,
    COLUMNS,
    FIRST_DATA_ROW,
    HEADER_ROW,
    LAST_DATA_ROW,
    SCAFFOLD_SLOTS,
    SHEET_NAME,
    TOTAL_DATA_ROWS,
    Role,
    ValueType,
    key_columns,
    match_key_columns,
    slotted_columns,
)

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "template" / "Format.xlsx"


@pytest.fixture(scope="module")
def ws():
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    return wb[SHEET_NAME]


def test_template_file_exists():
    assert TEMPLATE_PATH.exists(), f"template not found at {TEMPLATE_PATH}"


def test_sheet_name_matches(ws):
    assert ws.title == SHEET_NAME


def test_column_letters_are_a_through_s_in_order():
    assert COLUMN_LETTERS == tuple("ABCDEFGHIJKLMNOPQRS")
    assert len(COLUMNS) == 19


def test_column_index_matches_letter_position():
    for i, letter in enumerate(COLUMN_LETTERS, start=1):
        assert COLUMNS[letter].index == i


@pytest.mark.parametrize("letter", list("ABCDEFGHIJKLMNOPQRS"))
def test_header_matches_template_exactly(ws, letter):
    spec = COLUMNS[letter]
    actual = ws.cell(row=HEADER_ROW, column=spec.index).value
    assert actual == spec.header, f"column {letter}: template has {actual!r}, spec has {spec.header!r}"


def test_no_extra_populated_header_beyond_s(ws):
    extra = ws.cell(row=HEADER_ROW, column=20).value  # column T
    assert extra is None, f"template has an unmapped header at column T: {extra!r}"


def test_data_row_bounds(ws):
    last_with_content = 0
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is not None:
            last_with_content = r
    assert last_with_content == LAST_DATA_ROW
    assert TOTAL_DATA_ROWS == LAST_DATA_ROW - FIRST_DATA_ROW + 1 == 115


def test_first_data_row_has_content(ws):
    assert ws.cell(row=FIRST_DATA_ROW, column=1).value is not None


@pytest.mark.parametrize("letter", [c.letter for c in slotted_columns()])
def test_scaffold_matches_every_data_row(ws, letter):
    spec = COLUMNS[letter]
    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        actual = ws.cell(row=r, column=spec.index).value
        assert actual == spec.scaffold_raw, (
            f"column {letter} row {r}: scaffold mismatch\n  template: {actual!r}\n  spec:     {spec.scaffold_raw!r}"
        )


def test_slotted_columns_are_exactly_h_through_r():
    assert {c.letter for c in slotted_columns()} == set("HIJKLMNOPQR")


def test_anchor_column_is_slotted_and_agency_shaped():
    spec = COLUMNS[ANCHOR_COLUMN]
    assert spec.slotted
    assert "Agency" in spec.header or "agency" in spec.header.lower()


def test_non_slotted_columns_have_no_scaffold_raw():
    for c in COLUMNS.values():
        if not c.slotted:
            assert c.scaffold_raw is None, f"column {c.letter} is not slotted but has scaffold_raw set"


def test_key_columns_are_a_through_e():
    assert {c.letter for c in key_columns()} == set("ABCDE")


def test_match_keys_are_only_sno_and_plaza_name():
    assert {c.letter for c in match_key_columns()} == {"A", "C"}


def test_plaza_code_is_key_but_not_a_match_key():
    assert COLUMNS["B"].role == Role.KEY
    assert COLUMNS["B"].match_key is False


def test_village_and_remarks_are_optional():
    # F (Plaza Village/Location) was made optional at the user's request --
    # it's okay for a plaza to have no village data. S (Remarks) was
    # already the one genuinely optional field.
    optional = {c.letter for c in COLUMNS.values() if not c.required}
    assert optional == {"F", "S"}


def test_scaffold_slots_constant_matches_h_scaffold():
    # H's scaffold has exactly SCAFFOLD_SLOTS numbered markers.
    import re

    markers = re.findall(r"^\s*\d+\.", COLUMNS[ANCHOR_COLUMN].scaffold_raw, flags=re.MULTILINE)
    assert len(markers) == SCAFFOLD_SLOTS == 6


def test_composite_location_has_four_components():
    assert COLUMNS["F"].composite_components == ("village_name", "chainage", "city_name", "pincode")


def test_enum_column_has_values():
    # G is the only ENUM column left -- I moved to TEXT (presence only) at
    # the user's request, since real data sometimes has a contract-model
    # term there instead of EQ/Regular, and only slot count should matter.
    assert COLUMNS["G"].enum_values == ("Public Funded", "BOT", "TOT", "Invit", "MLFF")
    assert COLUMNS["I"].value_type == ValueType.TEXT
    assert COLUMNS["I"].enum_values == ()
