"""Phase 1 gate: merged-cell propagation, date normalization, and index lookup."""

from datetime import datetime

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

from app.excel_reader import build_index, read_cells


def _save(wb, tmp_path, name="test.xlsx") -> str:
    path = tmp_path / name
    wb.save(path)
    return str(path)


def test_merged_cell_propagates_text_and_anchor(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B2"] = "Contact Details"
    ws.merge_cells("B2:D2")
    path = _save(wb, tmp_path)

    records = read_cells(path)
    by_cell = {r.cell: r for r in records if r.sheet == "Sheet1"}

    assert by_cell["B2"].text == "Contact Details"
    assert by_cell["C2"].text == "Contact Details"
    assert by_cell["D2"].text == "Contact Details"

    assert by_cell["B2"].merge_anchor == "B2"
    assert by_cell["C2"].merge_anchor == "B2"
    assert by_cell["D2"].merge_anchor == "B2"

    assert by_cell["B2"].is_merged is True
    assert by_cell["C2"].is_merged is True


def test_date_cell_normalizes_and_preserves_raw_value(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = datetime(2026, 8, 10)
    ws["A1"].number_format = "dd/mm/yyyy"
    path = _save(wb, tmp_path)

    records = read_cells(path)
    a1 = next(r for r in records if r.cell == "A1")

    assert a1.text == "2026-08-10"
    assert isinstance(a1.raw_value, datetime)
    assert a1.raw_value.year == 2026
    assert a1.number_format == "dd/mm/yyyy"


def test_build_index_lookup(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    path = _save(wb, tmp_path)

    records = read_cells(path)
    index = build_index(records)

    assert index[("Sheet1", "A1")].text == "hello"
    assert ("Sheet1", "Z99") not in index


def test_comment_and_data_validation_captured(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Provide 10 numbers"
    ws["A1"].comment = Comment("please fill all 10", "author")

    dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(dv)
    dv.add("B1")
    ws["B1"] = "Yes"

    path = _save(wb, tmp_path)
    records = read_cells(path)
    by_cell = {r.cell: r for r in records}

    assert by_cell["A1"].has_comment is True
    assert by_cell["A1"].comment_text == "please fill all 10"
    assert by_cell["B1"].data_validation == '"Yes,No"'


def test_bold_flag(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "Not bold"
    path = _save(wb, tmp_path)

    records = read_cells(path)
    by_cell = {r.cell: r for r in records}
    assert by_cell["A1"].is_bold is True
    assert by_cell["A2"].is_bold is False


def test_blank_cell_has_empty_text(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "value"
    ws["A2"] = None  # blank, but not trailing -- A3 keeps it in the used range
    ws["A3"] = "another value"
    path = _save(wb, tmp_path)

    records = read_cells(path)
    a2 = next(r for r in records if r.cell == "A2")
    assert a2.text == ""
    assert a2.raw_value is None


def test_integral_float_normalizes_without_decimal(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10.0
    path = _save(wb, tmp_path)

    records = read_cells(path)
    a1 = next(r for r in records if r.cell == "A1")
    assert a1.text == "10"


def test_multiple_sheets_are_all_read(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "One"
    ws1["A1"] = "first"
    ws2 = wb.create_sheet("Two")
    ws2["A1"] = "second"
    path = _save(wb, tmp_path)

    records = read_cells(path)
    sheets = {r.sheet for r in records}
    assert sheets == {"One", "Two"}


def test_trailing_empty_rows_and_columns_are_trimmed(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "value"
    # Simulate an over-formatted template: a far-away cell touched by
    # formatting only, never given a value.
    ws["Z100"].number_format = "General"
    path = _save(wb, tmp_path)

    records = read_cells(path)
    coords = {r.cell for r in records}
    assert "Z100" not in coords


def test_internal_commas_and_slashes_survive_normalization(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "₹2,50,000"
    ws["A2"] = "10/08/2026"
    path = _save(wb, tmp_path)

    records = read_cells(path)
    by_cell = {r.cell: r for r in records}
    assert by_cell["A1"].text == "₹2,50,000"
    assert by_cell["A2"].text == "10/08/2026"
